import random
from time import sleep
import os

from selenium import webdriver
from openpyxl import load_workbook


class Bot:
    def __init__(self):
        self.driver = webdriver.Firefox()
        self.text_checker()

    def text_checker(self):
        this_folder = os.path.dirname(__file__)
        file = this_folder + '\\texts.xlsx'
        wb = load_workbook(file)
        sheet = wb.get_sheet_by_name('sheet1')
        row_count = sheet.max_row
        for x in range(2, row_count + 1):
            sleep(random.uniform(5, 7))
            text = sheet['A' + str(x)].value
            self.clean_text = text.replace('_x000D_', '')
            self.navigate()
            sheet['B' + str(x)].value = self.rating
            wb.save(file)
        self.driver.quit()

    def navigate(self):
        url = 'https://glvrd.ru/'
        self.driver.get(url)
        form = self.driver.find_element_by_class_name('ql-editor')
        form.clear()
        form.send_keys(self.clean_text)
        self.driver.implicitly_wait(5)
        self.rating = self.driver.find_element_by_xpath('//*[@id="app"]/div/div[3]/div[1]/div[2]/div[3]/div[1]/span[1]').text
        print(self.rating)
        return self.rating


if __name__ == "__main__":
    b = Bot()
import random
from time import sleep
import os

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook


class Bot:
    def __init__(self):
        self.driver = webdriver.Chrome()
        self.sign_in()

    def sign_in(self):
        url = 'https://turgenev.ashmanov.com/'
        self.driver.maximize_window()
        self.driver.get(url)
        sign_in_form = self.driver.find_element_by_css_selector('body > div.topblock > div.mb.large > div.last > span')
        sign_in_form.click()

        login_selector = '#dlg_main_block > form * input[name=email]'
        password_selector = '#dlg_main_block > form * input[name=password]'
        login = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, login_selector))
        )
        password = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, password_selector))
        )

        login.send_keys('')  # your login
        password.send_keys('')  # your password

        button_sign_in = self.driver.find_element_by_xpath('//*[@id="dlg_main_block"]/form/div[2]/input')
        sleep(random.uniform(3, 4))
        button_sign_in.click()

        self.editor_activation()

    def editor_activation(self):
        button_editor = WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.ID, 'show_tb_btn'))
        )
        button_editor.click()
        self.text_checker()

    def text_checker(self):
        this_folder = os.path.dirname(__file__)
        file = this_folder + '\\texts.xlsx'
        wb = load_workbook(file)
        sheet = wb.get_sheet_by_name('sheet1')
        row_count = sheet.max_row
        for x in range(2, row_count + 1):
            sleep(random.uniform(5, 7))
            text = sheet['A' + str(x)].value
            self.clean_text = text.replace('_x000D_', '')
            self.navigate()
            sheet['B' + str(x)].value = self.rating
            sheet['C' + str(x)].value = self.words_count + ' word(s)'
            sheet['D' + str(x)].value = self.text_size + ' characters'
            wb.save(file)
        self.driver.quit()
    
    def navigate(self):
        self.driver.switch_to.frame('textfield_ifr')
        form = self.driver.find_element_by_xpath('//*[@id="tinymce"]')
        form.clear()
        form.send_keys(self.clean_text)
        self.driver.switch_to.default_content()
        button_check_text = self.driver.find_element_by_xpath('//*[@id="recheck_btn"]')
        button_check_text.click()

        self.tabs = self.driver.window_handles
        if len(self.tabs) > 1:
            self.tab_manager()

        WebDriverWait(self.driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#infoblock > h2'))
        )

        self.rating = self.driver.find_element_by_css_selector('#infoblock > h2 > span > span.isum-block > span').text
        self.words_count = self.driver.find_element_by_id('words_count').text
        self.text_size = self.driver.find_element_by_id('text_size').text

        print(self.rating, self.words_count, self.text_size)

        return self.rating, self.words_count, self.text_size

    def tab_manager(self):
        self.driver.switch_to.window(self.tabs[-1])
        self.driver.close()
        self.driver.switch_to.window(self.tabs[0])


if __name__ == "__main__":
    b = Bot()
